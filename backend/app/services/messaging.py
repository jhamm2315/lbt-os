"""
Messaging service — channels, messages, file storage, AI assistant, XLSX export.
"""
from __future__ import annotations

import io
import re
import time
import uuid
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Any

from fastapi import HTTPException
from supabase import Client


# ---------------------------------------------------------------------------
# File type helpers
# ---------------------------------------------------------------------------

_EXT_TYPE_MAP = {
    "pdf": "pdf", "docx": "docx", "doc": "docx",
    "xlsx": "xlsx", "xls": "xlsx", "csv": "csv",
    "pptx": "pptx", "ppt": "pptx",
    "png": "image", "jpg": "image", "jpeg": "image",
    "gif": "image", "webp": "image",
    "txt": "text", "mp4": "video", "mov": "video",
}

_CHANNEL_TYPES = {"team", "ai_assistant", "announcements", "private"}

FILE_ICON = {
    "pdf": "📄", "docx": "📝", "xlsx": "📊", "pptx": "📑",
    "csv": "📋", "image": "🖼", "video": "🎬", "text": "📃", "other": "📎",
}

MAX_FILE_BYTES = 50 * 1024 * 1024  # 50 MB per file
STORAGE_BUCKET  = "message-files"

_ALLOWED_FILE_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "application/vnd.ms-powerpoint",
    "text/csv",
    "text/plain",
    "image/png", "image/jpeg", "image/gif", "image/webp",
    "video/mp4", "video/quicktime",
    "application/octet-stream",  # generic binary — allowed only for signed office formats
}

_EXT_MIME_MAP = {
    "pdf": {"application/pdf"},
    "docx": {"application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/octet-stream"},
    "doc": {"application/msword", "application/octet-stream"},
    "xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"},
    "xls": {"application/vnd.ms-excel", "application/octet-stream"},
    "csv": {"text/csv", "text/plain"},
    "pptx": {"application/vnd.openxmlformats-officedocument.presentationml.presentation", "application/octet-stream"},
    "ppt": {"application/vnd.ms-powerpoint", "application/octet-stream"},
    "png": {"image/png"},
    "jpg": {"image/jpeg"},
    "jpeg": {"image/jpeg"},
    "gif": {"image/gif"},
    "webp": {"image/webp"},
    "txt": {"text/plain"},
    "mp4": {"video/mp4"},
    "mov": {"video/quicktime"},
}

_ZIP_OFFICE_EXTS = {"docx", "xlsx", "pptx"}
_OLE_OFFICE_EXTS = {"doc", "xls", "ppt"}


# ---------------------------------------------------------------------------
# Channel operations
# ---------------------------------------------------------------------------

_DEFAULT_CHANNELS = [
    {"name": "answer-floor",   "channel_type": "team",         "description": "Ad hoc questions, handoffs, and fast operational answers"},
    {"name": "deal-desk",      "channel_type": "team",         "description": "Pipeline updates, quote notes, and customer context"},
    {"name": "answer-engine",  "channel_type": "ai_assistant", "description": "Call specialist bots for analytics on connected LBT OS data"},
    {"name": "wins-wire",      "channel_type": "announcements","description": "Announcements, launches, and customer wins"},
]


@dataclass(frozen=True)
class BusinessBot:
    key: str
    name: str
    title: str
    scope: str
    prompt: str
    examples: tuple[str, ...]


_BUSINESS_BOTS: dict[str, BusinessBot] = {
    "BI": BusinessBot(
        "BI", "BI", "Business Intelligence",
        "Executive KPI summaries, trend explanation, and decision support.",
        "You are BI, the Business Intelligence bot. Translate connected business data into executive-level answers with clear numbers, caveats, and recommended next moves.",
        ("@BI what changed in revenue this month?", "@BI summarize the health of the business."),
    ),
    "AE": BusinessBot(
        "AE", "AE", "Analytics Engineer",
        "Metric definitions, data joins, table logic, and reporting reliability.",
        "You are AE, the Analytics Engineer bot. Explain how metrics are derived, identify join/data-model risks, and suggest clean reporting logic.",
        ("@AE why do revenue and won leads not line up?", "@AE what tables support this answer?"),
    ),
    "DA": BusinessBot(
        "DA", "DA", "Data Analyst",
        "Ad hoc analysis, comparisons, segments, and fuzzy lookup across CRM records.",
        "You are DA, the Data Analyst bot. Answer ad hoc questions using the provided app data, compare segments, and call out where more data is needed.",
        ("@DA compare July sales to December.", "@DA which lead source is converting best?"),
    ),
    "DE": BusinessBot(
        "DE", "DE", "Data Engineer",
        "Data quality, sync health, missing fields, and pipeline readiness.",
        "You are DE, the Data Engineer bot. Focus on connected data health, missing fields, sync issues, schema gaps, and operational reliability.",
        ("@DE is our connected data complete enough?", "@DE what data is missing for staffing forecasts?"),
    ),
    "BDEV": BusinessBot(
        "BDEV", "BDEV", "Business Development",
        "Partnerships, lead sources, expansion, and growth opportunities.",
        "You are BDEV, the Business Development bot. Find growth angles, partnership opportunities, and source-level expansion plays using connected data only.",
        ("@BDEV where should we look for more deals?", "@BDEV which source deserves more outreach?"),
    ),
    "REVOPS": BusinessBot(
        "REVOPS", "REVOPS", "Revenue Operations",
        "Pipeline coverage, speed-to-lead, conversion, staffing, and capacity planning.",
        "You are REVOPS, the Revenue Operations bot. Connect marketing demand, pipeline capacity, follow-up speed, staffing, and revenue process recommendations.",
        ("@REVOPS how many support staff will we need for a busy rush?", "@REVOPS where is pipeline stuck?"),
    ),
    "M": BusinessBot(
        "M", "M", "Marketing",
        "Campaign performance, lead source mix, demand planning, and channel ROI.",
        "You are M, the Marketing bot. Evaluate campaign and source performance, lead quality, spend implications, and next best tests.",
        ("@M how are Google leads performing?", "@M should we increase Facebook ads?"),
    ),
    "S": BusinessBot(
        "S", "S", "Sales",
        "Win rate, deal follow-up, quote movement, and sales team priorities.",
        "You are S, the Sales bot. Focus on won/lost movement, follow-up urgency, deal quality, and practical sales actions.",
        ("@S who needs follow-up today?", "@S why is conversion down?"),
    ),
    "E": BusinessBot(
        "E", "E", "Expenses",
        "Spend mix, cost pressure, margin leakage, and vendor/category review.",
        "You are E, the Expenses bot. Analyze spend, margin pressure, recurring costs, vendors, and cost-control opportunities.",
        ("@E where are expenses hurting margin?", "@E compare marketing spend to revenue."),
    ),
    "HR": BusinessBot(
        "HR", "HR", "Human Resources",
        "Staffing estimates, workload coverage, hiring triggers, and schedule risk.",
        "You are HR, the Human Resources bot. Estimate staffing needs from demand, lead flow, revenue, and operational constraints, while clearly stating assumptions.",
        ("@HR how many staff do we need for the rush?", "@HR when should we hire support?"),
    ),
    "CS": BusinessBot(
        "CS", "CS", "Customer Success",
        "Retention, repeat customers, inactive accounts, and service recovery.",
        "You are CS, the Customer Success bot. Focus on retention, repeat purchase behavior, reactivation, and service follow-through.",
        ("@CS which customers are at risk?", "@CS how can we improve repeat rate?"),
    ),
    "OPS": BusinessBot(
        "OPS", "OPS", "Operations",
        "Capacity, workflow bottlenecks, job mix, and execution planning.",
        "You are OPS, the Operations bot. Translate data into staffing, scheduling, capacity, and process recommendations.",
        ("@OPS can we handle the next campaign spike?", "@OPS what bottleneck should we fix first?"),
    ),
    "FIN": BusinessBot(
        "FIN", "FIN", "Finance",
        "Profit, cash, margin, expense/revenue tradeoffs, and financial risk.",
        "You are FIN, the Finance bot. Focus on profit, margin, cash risk, spend discipline, and financial implications.",
        ("@FIN are we profitable enough?", "@FIN what is the biggest margin risk?"),
    ),
}

_BOT_ALIASES = {
    "MARKETING": "M",
    "SALES": "S",
    "EXPENSES": "E",
    "FINANCE": "FIN",
    "DATA": "DA",
    "ANALYST": "DA",
    "OPS": "OPS",
    "OPERATIONS": "OPS",
    "CUSTOMER": "CS",
    "SUCCESS": "CS",
}


def list_business_bots() -> list[dict[str, Any]]:
    return [
        {
            "key": bot.key,
            "name": bot.name,
            "title": bot.title,
            "scope": bot.scope,
            "examples": list(bot.examples),
        }
        for bot in _BUSINESS_BOTS.values()
    ]


def detect_bot_mentions(content: str) -> list[str]:
    found: list[str] = []
    for raw in re.findall(r"@([A-Za-z][A-Za-z0-9_-]{0,24})", content or ""):
        key = raw.upper().replace("-", "_")
        key = _BOT_ALIASES.get(key, key)
        if key in _BUSINESS_BOTS and key not in found:
            found.append(key)
    return found[:3]


def list_channels(db: Client, org_id: str, user_id: str | None = None) -> list[dict]:
    result = (
        db.table("message_channels")
        .select("*")
        .eq("org_id", org_id)
        .eq("is_archived", False)
        .order("created_at")
        .execute()
    )
    channels = result.data or []
    if not channels:
        channels = _seed_default_channels(db, org_id)
    if user_id:
        for channel in channels:
            _ensure_channel_member(db, org_id, channel["id"], user_id)
    return channels


def _seed_default_channels(db: Client, org_id: str) -> list[dict]:
    rows = [{**ch, "org_id": org_id} for ch in _DEFAULT_CHANNELS]
    result = db.table("message_channels").insert(rows).execute()
    return result.data or []


def create_channel(
    db: Client,
    org_id: str,
    created_by: str,
    name: str,
    channel_type: str = "team",
    description: str | None = None,
) -> dict:
    safe_name = _slug_channel_name(name)
    if not safe_name:
        raise HTTPException(400, "Channel name must contain letters or numbers.")
    if channel_type not in _CHANNEL_TYPES:
        raise HTTPException(400, "Unsupported channel type.")
    try:
        result = (
            db.table("message_channels")
            .insert({
                "org_id":       org_id,
                "name":         safe_name,
                "channel_type": channel_type,
                "description":  description,
                "created_by":   created_by,
            })
            .execute()
        )
        channel = result.data[0]
        _ensure_channel_member(db, org_id, channel["id"], created_by, role="owner")
        return channel
    except Exception as exc:
        if "unique" in str(exc).lower():
            raise HTTPException(400, f"A channel named '{safe_name}' already exists.")
        raise


# ---------------------------------------------------------------------------
# Message operations
# ---------------------------------------------------------------------------

def get_messages(
    db: Client,
    org_id: str,
    channel_id: str,
    limit: int = 50,
    before_id: str | None = None,
) -> dict[str, Any]:
    query = (
        db.table("messages")
        .select("*, message_files(*)")
        .eq("org_id", org_id)
        .eq("channel_id", channel_id)
        .order("created_at", desc=True)
        .limit(limit)
    )
    if before_id:
        cursor = (
            db.table("messages")
            .select("created_at")
            .eq("id", before_id)
            .eq("org_id", org_id)
            .maybe_single()
            .execute()
        )
        if cursor and cursor.data:
            query = query.lt("created_at", cursor.data["created_at"])

    result = query.execute()
    msgs = list(reversed(result.data or []))
    return {"messages": msgs, "has_more": len(result.data or []) == limit}


def send_message(
    db: Client,
    org_id: str,
    channel_id: str,
    sender_id: str,
    sender_name: str,
    content: str = "",
    message_type: str = "text",
    gif_url: str | None = None,
    reply_to_id: str | None = None,
    file_ids: list[str] | None = None,
    analytics: dict[str, Any] | None = None,
) -> dict:
    _ensure_channel_exists(db, org_id, channel_id)
    if message_type not in {"text", "gif", "ai_response", "analytics_response", "system"}:
        raise HTTPException(400, "Unsupported message type.")
    content = (content or "")[:4000]
    payload = {
        "org_id":       org_id,
        "channel_id":   channel_id,
        "sender_id":    sender_id,
        "sender_name":  sender_name,
        "content":      content,
        "message_type": message_type,
        "gif_url":      gif_url,
        "reply_to_id":  reply_to_id,
        "reactions":    {},
    }
    if analytics:
        payload["analytics"] = analytics
        message_type = "analytics_response" if message_type == "ai_response" else message_type
        payload["message_type"] = message_type

    try:
        result = db.table("messages").insert(payload).execute()
    except Exception as exc:
        if "analytics" not in str(exc).lower():
            raise
        # Allows local/dev installs to keep working before the analytics migration is applied.
        payload.pop("analytics", None)
        payload["message_type"] = "ai_response" if message_type == "analytics_response" else message_type
        result = db.table("messages").insert(payload).execute()
    message = result.data[0]
    _create_message_notifications(db, org_id, channel_id, message["id"], sender_id)

    # Link any pre-uploaded files to this message
    if file_ids:
        for fid in file_ids:
            db.table("message_files") \
              .update({"message_id": message["id"]}) \
              .eq("id", fid) \
              .eq("org_id", org_id) \
              .execute()

    files = (
        db.table("message_files")
        .select("*")
        .eq("message_id", message["id"])
        .execute()
    )
    message["message_files"] = files.data or []
    return message


def _ensure_channel_member(db: Client, org_id: str, channel_id: str, user_id: str, role: str = "member") -> None:
    try:
        db.table("message_channel_members").upsert({
            "org_id": org_id,
            "channel_id": channel_id,
            "user_id": user_id,
            "role": role,
        }, on_conflict="channel_id,user_id").execute()
    except Exception:
        pass


def _create_message_notifications(db: Client, org_id: str, channel_id: str, message_id: str, sender_id: str) -> None:
    try:
        members = (
            db.table("message_channel_members")
            .select("user_id")
            .eq("org_id", org_id)
            .eq("channel_id", channel_id)
            .neq("user_id", sender_id)
            .execute()
            .data or []
        )
        rows = [
            {
                "org_id": org_id,
                "channel_id": channel_id,
                "message_id": message_id,
                "user_id": member["user_id"],
                "status": "unread",
            }
            for member in members
            if member.get("user_id")
        ]
        if rows:
            db.table("message_notifications").upsert(rows, on_conflict="message_id,user_id").execute()
    except Exception:
        pass


def send_bot_responses_for_mentions(
    db: Client,
    org_id: str,
    channel_id: str,
    question: str,
    plan: str = "basic",
) -> list[dict]:
    bot_keys = detect_bot_mentions(question)
    responses = []
    for bot_key in bot_keys:
        responses.append(ask_business_bot_in_channel(db, org_id, channel_id, bot_key, question, plan=plan))
    return responses


def toggle_reaction(
    db: Client,
    org_id: str,
    message_id: str,
    user_id: str,
    emoji: str,
) -> dict:
    msg = (
        db.table("messages")
        .select("reactions")
        .eq("id", message_id)
        .eq("org_id", org_id)
        .maybe_single()
        .execute()
    )
    if not msg or not msg.data:
        raise HTTPException(404, "Message not found.")

    reactions: dict = dict(msg.data.get("reactions") or {})
    users: list = list(reactions.get(emoji, []))

    if user_id in users:
        users.remove(user_id)
    else:
        users.append(user_id)

    if users:
        reactions[emoji] = users
    else:
        reactions.pop(emoji, None)

    result = (
        db.table("messages")
        .update({"reactions": reactions})
        .eq("id", message_id)
        .eq("org_id", org_id)
        .execute()
    )
    return result.data[0]


# ---------------------------------------------------------------------------
# File upload / download
# ---------------------------------------------------------------------------

def _slug_channel_name(name: str) -> str:
    slug = re.sub(r"[^a-z0-9-]+", "-", (name or "").strip().lower())
    slug = re.sub(r"-{2,}", "-", slug).strip("-")
    return slug[:40]


def _safe_filename(filename: str) -> str:
    base = (filename or "attachment").rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    stem, dot, ext = base.rpartition(".")
    if not dot:
        stem, ext = base, ""
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", stem).strip(".-_")[:80] or "attachment"
    safe_ext = re.sub(r"[^A-Za-z0-9]+", "", ext.lower())[:12]
    return f"{safe_stem}.{safe_ext}" if safe_ext else safe_stem


def _ensure_channel_exists(db: Client, org_id: str, channel_id: str) -> None:
    channel = (
        db.table("message_channels")
        .select("id")
        .eq("id", channel_id)
        .eq("org_id", org_id)
        .maybe_single()
        .execute()
    )
    if not channel or not channel.data:
        raise HTTPException(404, "Channel not found.")


def _validate_file_upload(filename: str, ext: str, file_bytes: bytes, content_type: str) -> None:
    if not ext or ext not in _EXT_MIME_MAP:
        raise HTTPException(400, "Unsupported file extension.")
    if content_type not in _ALLOWED_FILE_TYPES or content_type not in _EXT_MIME_MAP[ext]:
        raise HTTPException(400, "File extension does not match its content type.")
    if not file_bytes:
        raise HTTPException(400, "File is empty.")

    header = file_bytes[:16]
    if ext == "pdf" and not header.startswith(b"%PDF-"):
        raise HTTPException(400, "PDF signature is invalid.")
    if ext == "png" and not header.startswith(b"\x89PNG\r\n\x1a\n"):
        raise HTTPException(400, "PNG signature is invalid.")
    if ext in {"jpg", "jpeg"} and not header.startswith(b"\xff\xd8\xff"):
        raise HTTPException(400, "JPEG signature is invalid.")
    if ext == "gif" and not (header.startswith(b"GIF87a") or header.startswith(b"GIF89a")):
        raise HTTPException(400, "GIF signature is invalid.")
    if ext == "webp" and not (header.startswith(b"RIFF") and file_bytes[8:12] == b"WEBP"):
        raise HTTPException(400, "WEBP signature is invalid.")
    if ext == "mp4" and b"ftyp" not in file_bytes[:32]:
        raise HTTPException(400, "MP4 signature is invalid.")
    if ext == "mov" and b"ftyp" not in file_bytes[:32]:
        raise HTTPException(400, "MOV signature is invalid.")
    if ext in _ZIP_OFFICE_EXTS and not header.startswith(b"PK\x03\x04"):
        raise HTTPException(400, "Office document signature is invalid.")
    if ext in _OLE_OFFICE_EXTS and not header.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        raise HTTPException(400, "Office document signature is invalid.")
    if ext in {"csv", "txt"} and b"\x00" in file_bytes[:4096]:
        raise HTTPException(400, "Text file appears to contain binary data.")

def upload_file(
    db: Client,
    org_id: str,
    channel_id: str,
    uploader_id: str,
    filename: str,
    file_bytes: bytes,
    content_type: str,
) -> dict:
    _ensure_channel_exists(db, org_id, channel_id)
    if len(file_bytes) > MAX_FILE_BYTES:
        raise HTTPException(413, "File exceeds the 50 MB limit.")

    safe_filename = _safe_filename(filename)
    ext = safe_filename.rsplit(".", 1)[-1].lower() if "." in safe_filename else ""
    file_type = _EXT_TYPE_MAP.get(ext, "other")

    _validate_file_upload(safe_filename, ext, file_bytes, content_type)

    storage_path = f"{org_id}/{channel_id}/{int(time.time())}_{uuid.uuid4().hex}_{safe_filename}"

    try:
        db.storage.from_(STORAGE_BUCKET).upload(
            path=storage_path,
            file=file_bytes,
            file_options={"content-type": content_type, "upsert": "false"},
        )
    except Exception as exc:
        raise HTTPException(500, "File storage failed. Check that the 'message-files' bucket exists in Supabase.")

    result = (
        db.table("message_files")
        .insert({
            "org_id":       org_id,
            "channel_id":   channel_id,
            "message_id":   None,
            "uploader_id":  uploader_id,
            "filename":     safe_filename,
            "file_type":    file_type,
            "content_type": content_type,
            "file_size":    len(file_bytes),
            "storage_path": storage_path,
        })
        .execute()
    )
    return result.data[0]


def get_file_signed_url(db: Client, org_id: str, file_id: str) -> dict:
    meta = (
        db.table("message_files")
        .select("*")
        .eq("id", file_id)
        .eq("org_id", org_id)
        .maybe_single()
        .execute()
    )
    if not meta or not meta.data:
        raise HTTPException(404, "File not found.")

    try:
        resp = db.storage.from_(STORAGE_BUCKET).create_signed_url(
            path=meta.data["storage_path"],
            expires_in=3600,
        )
        url = (
            resp.get("signedURL")
            or resp.get("data", {}).get("signedUrl")
            or ""
        )
    except Exception:
        url = ""

    return {
        "url":       url,
        "filename":  meta.data["filename"],
        "file_type": meta.data["file_type"],
        "file_size": meta.data["file_size"],
    }


# ---------------------------------------------------------------------------
# AI assistant in channel
# ---------------------------------------------------------------------------

def ask_ai_in_channel(
    db: Client,
    org_id: str,
    channel_id: str,
    sender_id: str,
    sender_name: str,
    question: str,
    plan: str = "basic",
) -> dict:
    from .metrics import get_dashboard_metrics, get_segment_analysis
    from .ai_audit import _get_llm_client_for_plan

    metrics  = get_dashboard_metrics(db, org_id, days=30)
    segments = get_segment_analysis(db, org_id, days=30)
    m = metrics

    seg_lines = [
        f"  {s['source']}: {s['conversion_rate_pct']:.1f}% conv, ${s['revenue']:,.0f} revenue"
        for s in (segments.get("segments") or [])[:5]
    ]

    system_prompt = (
        "You are the LBT-OS AI, a business intelligence assistant embedded in the team messaging system. "
        "Answer the question using the specific numbers provided. "
        "Be concise (3-5 sentences). Use real figures. If relevant, give one specific action."
    )

    user_prompt = f"""Business snapshot (last 30 days):
Revenue: ${m['revenue']['total']:,.2f}  |  Margin: {m['revenue']['margin_pct']}%  |  Avg deal: ${m['revenue']['avg_deal_size']:,.2f}
Leads: {m['leads']['total']} total  |  Won: {m['leads']['won']}  |  Conversion: {m['leads']['conversion_rate_pct']}%
Customers: {m['customers']['total']}  |  Repeat rate: {m['customers']['repeat_pct']}%
Expenses: ${m['expenses']['total']:,.2f}

Channel breakdown:
{chr(10).join(seg_lines) if seg_lines else '  (no channel data)'}

Question: {question}"""

    client, model, _ = _get_llm_client_for_plan(plan)
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_prompt},
        ],
        temperature=0.2,
        max_tokens=600,
    )
    answer = response.choices[0].message.content.strip()

    return send_message(
        db, org_id, channel_id,
        sender_id="ai_assistant",
        sender_name="LBT-OS AI",
        content=answer,
        message_type="ai_response",
    )


def ask_business_bot_in_channel(
    db: Client,
    org_id: str,
    channel_id: str,
    bot_key: str,
    question: str,
    plan: str = "basic",
) -> dict:
    from .ai_audit import _get_llm_client_for_plan

    bot = _BUSINESS_BOTS.get(bot_key.upper())
    if not bot:
        raise HTTPException(404, "Business bot not found.")

    context = _build_connected_data_context(db, org_id, question)
    system_prompt = (
        f"{bot.prompt}\n\n"
        "Hard rule: use only the connected LBT OS application data provided in the prompt. "
        "Do not claim access to external systems, the web, unsupported HR systems, or unseen ad platforms. "
        "If the user asks for something that needs unavailable data, state exactly what connected data is missing and give a practical assumption-based estimate only when reasonable. "
        "Keep the answer concise, operational, and specific. Include numbers when provided."
    )
    user_prompt = f"""Bot: @{bot.key} — {bot.title}
Question from channel:
{question}

Connected LBT OS data context:
{context}

Answer format:
- Direct answer first.
- Then 2-4 bullets with supporting numbers or caveats.
- End with one next action."""

    client, model, _ = _get_llm_client_for_plan(plan)
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=0.15,
        max_tokens=750,
    )
    answer = response.choices[0].message.content.strip()
    analytics = _build_analytics_payload(db, org_id, bot.key, question)

    return send_message(
        db, org_id, channel_id,
        sender_id=f"bot_{bot.key.lower()}",
        sender_name=f"{bot.key} · {bot.title}",
        content=answer,
        message_type="ai_response",
        analytics=analytics,
    )


def _build_analytics_payload(db: Client, org_id: str, bot_key: str, question: str) -> dict[str, Any]:
    from .metrics import get_dashboard_metrics, get_revenue_forecast, get_segment_analysis
    from .revenue_intelligence import get_data_quality_scorecard, get_speed_to_lead, get_stage_aging, get_win_loss_cohort

    charts: list[dict[str, Any]] = []
    notes: list[str] = []

    sales_rows = _safe_table_rows(
        db,
        "sales",
        "service, amount, cost, profit, payment_status, source, sold_at",
        org_id,
        limit=500,
        order="sold_at",
    )
    expense_rows = _safe_table_rows(
        db,
        "expenses",
        "category, description, amount, vendor, is_recurring, expense_date",
        org_id,
        limit=500,
        order="expense_date",
    )
    lead_rows = _safe_table_rows(
        db,
        "leads",
        "source, status, estimated_value, follow_up_at, contacted_at, created_at",
        org_id,
        limit=500,
        order="created_at",
    )

    paid_sales = [row for row in sales_rows if row.get("payment_status") == "paid"]
    monthly_revenue = _group_sales_by_month(paid_sales)
    if monthly_revenue:
        charts.append({
            "id": "monthly_revenue",
            "type": "line",
            "title": "Revenue Trend",
            "subtitle": "Paid sales grouped by month from connected sales records.",
            "data": monthly_revenue,
            "xKey": "month",
            "series": [{"key": "revenue", "name": "Revenue", "format": "currency"}],
        })

    try:
        segments = get_segment_analysis(db, org_id, days=365)
    except Exception:
        segments = {}
    source_data = [
        {
            "source": str(seg.get("source") or "unknown").replace("_", " "),
            "revenue": round(float(seg.get("revenue") or 0), 2),
            "leads": int(seg.get("leads") or 0),
            "conversion": round(float(seg.get("conversion_rate_pct") or 0), 1),
        }
        for seg in (segments.get("segments") or [])[:8]
    ]
    if source_data:
        charts.append({
            "id": "source_performance",
            "type": "bar",
            "title": "Source Performance",
            "subtitle": "Revenue, lead volume, and conversion by connected lead source.",
            "data": source_data,
            "xKey": "source",
            "series": [
                {"key": "revenue", "name": "Revenue", "format": "currency"},
                {"key": "leads", "name": "Leads", "format": "number"},
            ],
        })

    expense_data = _group_amounts(expense_rows, "category", "amount", label_key="category")
    if expense_data:
        charts.append({
            "id": "expense_mix",
            "type": "pie",
            "title": "Expense Mix",
            "subtitle": "Spend by category from connected expense records.",
            "data": expense_data[:8],
            "nameKey": "category",
            "valueKey": "amount",
            "format": "currency",
        })

    pipeline_data = _group_counts(lead_rows, "status", label_key="stage")
    if pipeline_data:
        charts.append({
            "id": "pipeline_stage_count",
            "type": "bar",
            "title": "Pipeline Shape",
            "subtitle": "Lead counts by current CRM stage.",
            "data": pipeline_data,
            "xKey": "stage",
            "series": [{"key": "count", "name": "Leads", "format": "number"}],
        })

    key = bot_key.upper()
    lower_q = question.lower()
    if key in {"REVOPS", "HR", "OPS", "S"} or any(term in lower_q for term in ("staff", "rush", "capacity", "pipeline", "follow")):
        try:
            speed = get_speed_to_lead(db, org_id, days=90)
            speed_data = [
                {"source": row.get("source"), "hours": row.get("avg_hours"), "sample": row.get("sample_size")}
                for row in speed.get("by_source", [])[:8]
            ]
            if speed_data:
                charts.append({
                    "id": "speed_to_lead",
                    "type": "bar",
                    "title": "Speed To Lead",
                    "subtitle": "Average contact delay by source. Lower is better.",
                    "data": speed_data,
                    "xKey": "source",
                    "series": [{"key": "hours", "name": "Avg Hours", "format": "number"}],
                })
        except Exception:
            pass
        try:
            aging = get_stage_aging(db, org_id)
            aging_data = [
                {"stage": row.get("stage"), "count": row.get("count"), "avg_days": row.get("avg_days_in_stage")}
                for row in aging.get("stages", [])[:8]
            ]
            if aging_data:
                charts.append({
                    "id": "stage_aging",
                    "type": "bar",
                    "title": "Stage Aging",
                    "subtitle": "Open lead age by stage from CRM stage timestamps.",
                    "data": aging_data,
                    "xKey": "stage",
                    "series": [
                        {"key": "count", "name": "Open Leads", "format": "number"},
                        {"key": "avg_days", "name": "Avg Days", "format": "number"},
                    ],
                })
        except Exception:
            pass

    if key in {"BI", "DA", "FIN"} or any(term in lower_q for term in ("forecast", "yoy", "compare", "july", "dec")):
        yoy_data = _monthly_yoy_sales(paid_sales)
        if yoy_data:
            charts.append({
                "id": "monthly_yoy",
                "type": "bar",
                "title": "Monthly Revenue Comparison",
                "subtitle": "Paid revenue by calendar month and year where connected data exists.",
                "data": yoy_data,
                "xKey": "month",
                "series": _yoy_series(yoy_data),
            })
        try:
            forecast = get_revenue_forecast(db, org_id, lookback_weeks=16)
            summary = forecast.get("summary") or {}
            forecast_data = [
                {"period": "30 days", "revenue": summary.get("next_30_days") or 0},
                {"period": "60 days", "revenue": summary.get("next_60_days") or 0},
                {"period": "90 days", "revenue": summary.get("next_90_days") or 0},
            ]
            if any(row["revenue"] for row in forecast_data):
                charts.append({
                    "id": "forecast_summary",
                    "type": "bar",
                    "title": "Revenue Forecast",
                    "subtitle": forecast.get("narrative") or "Linear projection from recent connected sales.",
                    "data": forecast_data,
                    "xKey": "period",
                    "series": [{"key": "revenue", "name": "Projected Revenue", "format": "currency"}],
                })
        except Exception:
            pass

    if key in {"DE", "AE"}:
        try:
            quality = get_data_quality_scorecard(db, org_id)
            quality_data = [
                {"field": f"{row.get('entity')}.{row.get('field')}", "complete": row.get("pct") or 0}
                for row in quality.get("fields", [])[:10]
            ]
            if quality_data:
                charts.append({
                    "id": "data_quality",
                    "type": "bar",
                    "title": "Data Completeness",
                    "subtitle": f"Overall data quality grade: {quality.get('grade', 'n/a')}.",
                    "data": quality_data,
                    "xKey": "field",
                    "series": [{"key": "complete", "name": "% Complete", "format": "percent"}],
                })
        except Exception:
            pass

    if not charts:
        notes.append("No chartable connected data was available for this request yet.")

    return {
        "version": 1,
        "source": "connected_lbt_os_data",
        "bot": bot_key.upper(),
        "question": question,
        "charts": charts[:5],
        "notes": notes,
    }


def _group_sales_by_month(sales_rows: list[dict]) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, float]] = {}
    for row in sales_rows:
        sold_at = str(row.get("sold_at") or "")
        month = sold_at[:7]
        if len(month) != 7:
            continue
        bucket = buckets.setdefault(month, {"month": month, "revenue": 0.0, "profit": 0.0})
        bucket["revenue"] += float(row.get("amount") or 0)
        bucket["profit"] += float(row.get("profit") or 0)
    return [
        {"month": month, "revenue": round(data["revenue"], 2), "profit": round(data["profit"], 2)}
        for month, data in sorted(buckets.items())[-12:]
    ]


def _group_amounts(rows: list[dict], group_key: str, amount_key: str, label_key: str = "label") -> list[dict[str, Any]]:
    buckets: dict[str, float] = {}
    for row in rows:
        label = str(row.get(group_key) or "unknown").replace("_", " ")
        buckets[label] = buckets.get(label, 0.0) + float(row.get(amount_key) or 0)
    return [
        {label_key: label, "amount": round(amount, 2)}
        for label, amount in sorted(buckets.items(), key=lambda item: item[1], reverse=True)
    ]


def _group_counts(rows: list[dict], group_key: str, label_key: str = "label") -> list[dict[str, Any]]:
    buckets: dict[str, int] = {}
    for row in rows:
        label = str(row.get(group_key) or "unknown").replace("_", " ")
        buckets[label] = buckets.get(label, 0) + 1
    return [
        {label_key: label, "count": count}
        for label, count in sorted(buckets.items(), key=lambda item: item[1], reverse=True)
    ]


def _monthly_yoy_sales(sales_rows: list[dict]) -> list[dict[str, Any]]:
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    years = sorted({str(row.get("sold_at") or "")[:4] for row in sales_rows if len(str(row.get("sold_at") or "")) >= 7})[-3:]
    if not years:
        return []
    data = [{"month": name, **{year: 0.0 for year in years}} for name in month_names]
    for row in sales_rows:
        sold_at = str(row.get("sold_at") or "")
        if len(sold_at) < 7:
            continue
        year = sold_at[:4]
        if year not in years:
            continue
        try:
            month_idx = int(sold_at[5:7]) - 1
        except ValueError:
            continue
        if 0 <= month_idx < 12:
            data[month_idx][year] += float(row.get("amount") or 0)
    return [
        {key: round(value, 2) if isinstance(value, float) else value for key, value in row.items()}
        for row in data
        if any(row.get(year, 0) for year in years)
    ]


def _yoy_series(yoy_data: list[dict]) -> list[dict[str, str]]:
    years = [key for key in yoy_data[0].keys() if key != "month"] if yoy_data else []
    return [{"key": year, "name": year, "format": "currency"} for year in years]


def _safe_table_rows(db: Client, table: str, select: str, org_id: str, limit: int = 100, order: str | None = None) -> list[dict]:
    try:
        query = db.table(table).select(select).eq("org_id", org_id).limit(limit)
        if order:
            query = query.order(order, desc=True)
        return query.execute().data or []
    except Exception:
        return []


def _money(value: Any) -> str:
    try:
        return f"${float(value or 0):,.0f}"
    except Exception:
        return "$0"


def _build_connected_data_context(db: Client, org_id: str, question: str) -> str:
    from .metrics import get_dashboard_metrics, get_revenue_forecast, get_segment_analysis
    from .revenue_intelligence import (
        get_data_quality_scorecard,
        get_expansion_signals,
        get_speed_to_lead,
        get_stage_aging,
        get_win_loss_cohort,
    )

    try:
        metrics = get_dashboard_metrics(db, org_id, days=365)
    except Exception:
        metrics = {}
    try:
        current = get_dashboard_metrics(db, org_id, days=30)
    except Exception:
        current = {}
    try:
        segments = get_segment_analysis(db, org_id, days=365)
    except Exception:
        segments = {}

    revenue = metrics.get("revenue", {})
    leads = metrics.get("leads", {})
    customers = metrics.get("customers", {})
    expenses = metrics.get("expenses", {})
    current_revenue = (current.get("revenue") or {}).get("total", 0)

    leads_rows = _safe_table_rows(
        db,
        "leads",
        "id, name, source, status, service_interest, estimated_value, notes, assigned_to, follow_up_at, contacted_at, created_at",
        org_id,
        limit=150,
        order="created_at",
    )
    sales_rows = _safe_table_rows(
        db,
        "sales",
        "id, service, amount, cost, profit, payment_status, source, sold_at, notes",
        org_id,
        limit=150,
        order="sold_at",
    )
    customers_rows = _safe_table_rows(
        db,
        "customers",
        "id, name, email, tags, lifetime_value, total_orders, last_purchase_at, notes, created_at",
        org_id,
        limit=120,
        order="created_at",
    )
    expense_rows = _safe_table_rows(
        db,
        "expenses",
        "id, category, description, amount, vendor, is_recurring, recurrence_period, expense_date",
        org_id,
        limit=150,
        order="expense_date",
    )
    integration_rows = _safe_table_rows(
        db,
        "integration_connections",
        "provider, label, status, external_account_name, last_synced_at, last_sync_status, last_sync_error",
        org_id,
        limit=25,
        order="updated_at",
    )

    fuzzy_matches = _fuzzy_matches(question, {
        "leads": leads_rows,
        "sales": sales_rows,
        "customers": customers_rows,
        "expenses": expense_rows,
        "integrations": integration_rows,
    })

    try:
        win_loss = get_win_loss_cohort(db, org_id, days=365)
    except Exception:
        win_loss = {}
    try:
        speed = get_speed_to_lead(db, org_id, days=90)
    except Exception:
        speed = {}
    try:
        quality = get_data_quality_scorecard(db, org_id)
    except Exception:
        quality = {}
    try:
        expansion = get_expansion_signals(db, org_id)
    except Exception:
        expansion = {}
    try:
        aging = get_stage_aging(db, org_id)
    except Exception:
        aging = {}
    try:
        forecast = get_revenue_forecast(db, org_id, lookback_weeks=16)
    except Exception:
        forecast = {}

    segment_lines = [
        f"- {seg.get('source')}: {seg.get('leads')} leads, {seg.get('won')} won, {seg.get('conversion_rate_pct')}% conversion, {_money(seg.get('revenue'))} revenue"
        for seg in (segments.get("segments") or [])[:8]
    ]
    integration_lines = [
        f"- {row.get('provider')}: {row.get('status')} / last sync {row.get('last_sync_status') or 'unknown'} at {row.get('last_synced_at') or 'never'}"
        for row in integration_rows[:8]
    ]
    fuzzy_lines = [
        f"- {match['table']}: {match['summary']}"
        for match in fuzzy_matches[:12]
    ]

    return f"""Available connected data only:
- Leads: {leads.get('total', 0)} total, {leads.get('won', 0)} won, {leads.get('lost', 0)} lost, {leads.get('conversion_rate_pct', 0)}% conversion, {leads.get('missed_follow_ups', 0)} missed follow-ups.
- Revenue: {_money(revenue.get('total'))} last 365 days, {_money(current_revenue)} last 30 days, {_money(revenue.get('profit'))} gross profit, {revenue.get('margin_pct', 0)}% margin, {_money(revenue.get('avg_deal_size'))} average deal.
- Customers: {customers.get('total', 0)} total, {customers.get('repeat', 0)} repeat, {customers.get('repeat_pct', 0)}% repeat rate.
- Expenses: {_money(expenses.get('total'))} total; categories: {expenses.get('by_category', {})}.
- Connected sources: {integration_lines if integration_lines else 'No connector rows found.'}

Channel/source analysis:
{chr(10).join(segment_lines) if segment_lines else '- No source segment analysis available.'}

Revenue intelligence:
- Win/loss cohorts: {win_loss.get('cohorts', [])[:8]}
- Speed to lead: {speed}
- Stage aging: {aging}
- Data quality: {quality}
- Expansion/reactivation signals: {(expansion.get('signals') or [])[:8]}
- Revenue forecast: {forecast}

Fuzzy matches from connected app records:
{chr(10).join(fuzzy_lines) if fuzzy_lines else '- No close record matches found for the question terms.'}
"""


def _fuzzy_matches(question: str, tables: dict[str, list[dict]]) -> list[dict[str, str]]:
    query = re.sub(r"@\w+", " ", question or "").strip().lower()
    if not query:
        return []

    matches: list[dict[str, Any]] = []
    for table, rows in tables.items():
        for row in rows:
            parts = []
            for value in row.values():
                if value is None:
                    continue
                if isinstance(value, (str, int, float, bool)):
                    parts.append(str(value))
                elif isinstance(value, list):
                    parts.extend(str(item) for item in value)
            haystack = " ".join(parts).lower()
            if not haystack:
                continue
            token_hits = sum(1 for token in re.findall(r"[a-z0-9]{3,}", query) if token in haystack)
            ratio = SequenceMatcher(None, query[:240], haystack[:500]).ratio()
            score = ratio + (token_hits * 0.18)
            if score >= 0.22 or token_hits:
                matches.append({
                    "score": score,
                    "table": table,
                    "summary": _summarize_row(row),
                })

    matches.sort(key=lambda item: item["score"], reverse=True)
    return [{"table": item["table"], "summary": item["summary"]} for item in matches[:15]]


def _summarize_row(row: dict) -> str:
    preferred = [
        "name", "service", "description", "source", "status", "amount", "estimated_value",
        "category", "vendor", "payment_status", "sold_at", "expense_date", "created_at",
        "last_purchase_at", "notes",
    ]
    parts = []
    for key in preferred:
        value = row.get(key)
        if value not in (None, "", []):
            parts.append(f"{key}={value}")
    if not parts:
        parts = [f"{key}={value}" for key, value in row.items() if value not in (None, "", [])][:6]
    return "; ".join(parts)[:500]


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def export_channel_xlsx(db: Client, org_id: str, channel_id: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    channel = (
        db.table("message_channels")
        .select("name")
        .eq("id", channel_id)
        .eq("org_id", org_id)
        .maybe_single()
        .execute()
    )
    channel_name = (channel.data or {}).get("name", "channel")

    result = (
        db.table("messages")
        .select("*, message_files(filename, file_type, file_size)")
        .eq("org_id", org_id)
        .eq("channel_id", channel_id)
        .order("created_at")
        .limit(2000)
        .execute()
    )
    msgs = result.data or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = channel_name[:31]

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["Timestamp", "Sender", "Type", "Message / Content", "Attachments"]
    col_widths = [22, 22, 14, 80, 40]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False, vertical="center")
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.row_dimensions[1].height = 24

    for row_idx, msg in enumerate(msgs, 2):
        files = msg.get("message_files") or []
        file_names = ", ".join(f.get("filename", "") for f in files)
        content = msg.get("content") or msg.get("gif_url") or ""

        ws.cell(row=row_idx, column=1, value=(msg.get("created_at") or "")[:19])
        ws.cell(row=row_idx, column=2, value=msg.get("sender_name", ""))
        ws.cell(row=row_idx, column=3, value=msg.get("message_type", "text"))
        cell = ws.cell(row=row_idx, column=4, value=content)
        cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=5, value=file_names)
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
